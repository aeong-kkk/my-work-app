"""기능 1 -- 발주 엑셀에서 정보 추출 + 리드타임 계산 (feature-1-spec.md 구현)

입력: 발주 엑셀 파일 1개 (Info 시트 1개뿐 -- 예전 Packing 시트는 Info!I1 "Shipping Information" 구역으로 통합됨)
동작: feature-1-spec.md '동작' 단계 참고
출력: 결과 행(들) -- 발주일/구분/요청일/모델명/내부명/수량/층구성/리드타임(일수)/비고
"""

import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

REQUIRED_SHEETS = ("Info",)
SHIPPING_INFO_START_ROW = 3
SHIPPING_QTY_COL = "I"
SHIPPING_DATE_COL = "M"


class OrderFileError(Exception):
    """Info/Packing 시트가 없거나 구조가 예상과 다를 때"""


def _is_blank(value):
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _clean_text(value):
    if value is None:
        return None
    return str(value).strip()


def _classify_order_type(value):
    """Info F3 ("NEW／Repeat") 값을 신규발주/재발주로 정규화한다."""
    text = _clean_text(value)
    if text is None:
        return None
    lowered = text.lower()
    if "new" in lowered:
        return "신규발주"
    if "repeat" in lowered:
        return "재발주"
    return text  # 인식 못한 값은 원문 그대로 반환


def _to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return value


def _correct_year(value, order_date):
    """'월/일'만 있는 문자열 요청일에 연도를 보정한다. 이미 날짜면 그대로 반환."""
    d = _to_date(value)
    if isinstance(d, date):
        return d
    if d is None:
        return None
    match = re.search(r"(\d{1,2})\s*/\s*(\d{1,2})", str(d))
    if not match:
        return None
    month, day = int(match.group(1)), int(match.group(2))
    year = order_date.year if order_date else date.today().year
    return date(year, month, day)


def _extract_quantity(value):
    """수량 값에서 숫자만 분리한다 (예: '16\\n(IST Coupon)' -> 16)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"\d+", str(value))
    return int(match.group()) if match else None


def _filled_rows(ws, start_row, col):
    rows = []
    row = start_row
    while True:
        cell = ws[f"{col}{row}"]
        if _is_blank(cell.value):
            break
        rows.append(row)
        row += 1
    return rows


def extract_order_records(file_path):
    """발주 엑셀 파일 1개를 읽어 결과 행(들)을 반환한다."""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 1. Info 시트 존재 확인
    missing_sheets = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing_sheets:
        raise OrderFileError(f"필수 시트 없음: {', '.join(missing_sheets)}")

    info = wb["Info"]

    # 2~6. 발주일 / 구분(신규발주·재발주) / 모델명(Customer Model Name) / 내부명(Internal Model Name) / Total-T(板厚) / 층구성(层构成)
    order_date = _to_date(info["G1"].value)
    order_type = _classify_order_type(info["F3"].value)
    model_name = _clean_text(info["F5"].value)
    internal_name = _clean_text(info["F6"].value)
    total_t = _clean_text(info["G18"].value)
    layer_config = _clean_text(info["G19"].value)

    # 6. "Shipping Information" 구역(I열 기준)에 값이 입력된 행 확인
    shipping_rows = _filled_rows(info, SHIPPING_INFO_START_ROW, SHIPPING_QTY_COL)
    if not shipping_rows:
        raise OrderFileError("Shipping Information 구역(I열)에 값이 없음")

    # 7. 각 행 -> 수량(I열)/요청일(M열), 행 개수만큼 결과 생성
    raw_records = [
        {
            "quantity": info[f"{SHIPPING_QTY_COL}{row}"].value,
            "request_date": info[f"{SHIPPING_DATE_COL}{row}"].value,
        }
        for row in shipping_rows
    ]

    # 8. 요청일 연도 보정 (모든 행 먼저 계산 -- 아래 "분할입고" 판정에 필요)
    request_dates = [_correct_year(raw["request_date"], order_date) for raw in raw_records]

    # 9. 행이 여러 개면서 요청일이 서로 다르면, 그중 가장 늦은 요청일인 행은 구분을 "분할입고"로 표시
    distinct_dates = {d for d in request_dates if d is not None}
    split_arrival_date = max(distinct_dates) if len(raw_records) > 1 and len(distinct_dates) > 1 else None

    results = []
    for raw, request_date in zip(raw_records, request_dates):
        # 10. 수량 값에서 숫자만 분리
        quantity = _extract_quantity(raw["quantity"])
        # 11. 값 앞뒤 공백/줄바꿈 제거 -- 모델명/층구성은 _clean_text에서 이미 적용

        # 12. 발주일/요청일로 리드타임(일수) 계산 = 요청일 - 발주일
        lead_time = (request_date - order_date).days if order_date and request_date else None

        row_order_type = "분할입고" if split_arrival_date is not None and request_date == split_arrival_date else order_type

        # 13. 필수 항목 누락 확인 -> "누락" 플래그
        field_values = {
            "발주일": order_date,
            "요청일": request_date,
            "구분": row_order_type,
            "모델명": model_name,
            "내부명": internal_name,
            "Total-T": total_t,
            "수량": quantity,
            "층구성": layer_config,
        }
        missing_fields = [name for name, val in field_values.items() if _is_blank(val)]
        note = f"누락({','.join(missing_fields)})" if missing_fields else ""

        # 14. 정리된 값들을 결과 행으로 묶기
        results.append({
            "발주일": order_date,
            "구분": row_order_type,
            "요청일": request_date,
            "모델명": model_name,
            "내부명": internal_name,
            "Total-T": total_t,
            "수량": quantity,
            "층구성": layer_config,
            "리드타임(일수)": lead_time,
            "비고": note,
        })

    return results


def _format_value(value):
    if value is None:
        return ""
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def print_results(records):
    columns = ["발주일", "구분", "요청일", "모델명", "내부명", "Total-T", "수량", "층구성", "리드타임(일수)", "비고"]
    widths = {
        c: max([len(c)] + [len(_format_value(r[c])) for r in records])
        for c in columns
    }

    header = " | ".join(c.ljust(widths[c]) for c in columns)
    print(header)
    print("-+-".join("-" * widths[c] for c in columns))
    for r in records:
        print(" | ".join(_format_value(r[c]).ljust(widths[c]) for c in columns))


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else "inputs/sample-order-input.xlsx"
    path = Path(target)
    if not path.is_absolute():
        path = Path(__file__).parent / target

    try:
        records = extract_order_records(path)
    except OrderFileError as e:
        print(f"에러: {e}")
        sys.exit(1)

    print(f"[입력 파일] {path.name}\n")
    print_results(records)
