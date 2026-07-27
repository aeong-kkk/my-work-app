"""기능 8 -- 발주 정보로 견적서(고객제출용) + 발주서(FPE PO, 내부보관용) 자동 작성 (feature-8-spec.md 구현)

입력: 발주 엑셀 파일 1개(기능 1과 동일 형식), TYPE, 제품 개취수, 툴비 포함 여부, 견적서 양식 파일
동작: feature-8-spec.md '동작' 단계 참고
출력: outputs/ 폴더에 완성된 견적서 파일(.xlsx) -- 저장된 경로 반환
"""

import re
import sys
from pathlib import Path

import openpyxl

from feature1_extract_order import OrderFileError, extract_order_records

DEFAULT_TEMPLATE = Path(__file__).parent / "inputs" / "견적서작성 양식.xlsx"
OUTPUTS_DIR = Path(__file__).parent / "outputs"

PRICE_DB_SHEET = "PRICE_DB"
CUSTOMER_SHEET = "Customer "
FPE_PO_SHEET = "FPE PO "
REQUIRED_SHEETS = (PRICE_DB_SHEET, CUSTOMER_SHEET, FPE_PO_SHEET)

LAYER_CELL = "I3"
TYPE_CELL = "I4"
COUNT_CELL = "I5"
FPE_MARGIN_CELL = "I9"  # PRICE_DB에 이미 있는 "FPE 단가"(=고객 단가/1.035) 수식

MODEL_CELL = "B16"
LAYER_TEXT_CELL = "B17"
QTY_CELL = "G16"
TOOL_QTY_CELL = "G20"

FPE_UNIT_PRICE_CELL = "I21"


class QuoteInputError(Exception):
    """층구성 파싱 실패 등 견적서 작성에 필요한 입력이 잘못됐을 때"""


def _extract_layer_count(layer_config):
    """층구성 문자열(예: '10L/FVSS3')에서 'L' 앞 숫자만 추출한다."""
    if not layer_config:
        raise QuoteInputError("층구성 값이 없어 층수를 알 수 없습니다.")
    match = re.search(r"(\d+)\s*[Ll]", layer_config)
    if not match:
        raise QuoteInputError(f"층구성 값에서 층수를 못 찾았습니다: {layer_config!r}")
    return int(match.group(1))


def _validate_layer_type(template_path, layer_count, product_type):
    """PRICE_DB 원본표에 층수+TYPE 조합이 실제로 있는지 확인한다.

    이 표는 층수마다 쓸 수 있는 TYPE이 정해져 있어서(예: 층 2는 'Interposer'만,
    층 10은 'C'/'D'/'3D' 등만), 잘못된 조합을 넣으면 엑셀에서 단가가 조용히
    빈 값/오류로 나온다. 저장 전에 미리 걸러서 알려준다.
    """
    wb = openpyxl.load_workbook(template_path, data_only=True)
    price_db = wb[PRICE_DB_SHEET]

    types_for_layer = set()
    found = False
    for row in price_db.iter_rows(min_row=4, max_col=2):
        layer_val, type_val = (cell.value for cell in row)
        if layer_val is None:
            continue
        if layer_val == layer_count:
            types_for_layer.add(type_val)
            if type_val == product_type:
                found = True

    if not types_for_layer:
        raise QuoteInputError(f"PRICE_DB에 층수 {layer_count}에 해당하는 데이터가 없습니다.")
    if not found:
        options = ", ".join(sorted(t for t in types_for_layer if t is not None))
        raise QuoteInputError(
            f"층수 {layer_count}에는 TYPE {product_type!r}가 없습니다. "
            f"이 층수에서 쓸 수 있는 TYPE: {options}"
        )


def _safe_filename_part(text):
    return re.sub(r'[\\/:*?"<>|]', "_", str(text)).strip() or "unknown"


def generate_quote(
    order_file,
    product_type,
    product_count,
    include_tool_charge=True,
    template_path=DEFAULT_TEMPLATE,
    output_dir=OUTPUTS_DIR,
):
    """발주 엑셀 + 사람 입력값으로 견적서 파일을 만들어 저장하고, 저장된 경로를 반환한다."""
    records = extract_order_records(order_file)

    model_name = records[0]["모델명"]
    layer_config = records[0]["층구성"]
    order_date = records[0]["발주일"]
    total_quantity = sum(r["수량"] for r in records if r["수량"] is not None)

    layer_count = _extract_layer_count(layer_config)
    _validate_layer_type(template_path, layer_count, product_type)

    wb = openpyxl.load_workbook(template_path, data_only=False)
    missing_sheets = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing_sheets:
        raise QuoteInputError(f"견적서 양식에 필수 시트 없음: {', '.join(missing_sheets)}")

    price_db = wb[PRICE_DB_SHEET]
    customer = wb[CUSTOMER_SHEET]
    fpe_po = wb[FPE_PO_SHEET]

    price_db[LAYER_CELL] = layer_count
    price_db[TYPE_CELL] = product_type
    price_db[COUNT_CELL] = product_count

    customer[MODEL_CELL] = model_name
    customer[LAYER_TEXT_CELL] = layer_config
    customer[QTY_CELL] = total_quantity
    customer[TOOL_QTY_CELL] = 1 if include_tool_charge else 0

    fpe_price_cell = fpe_po[FPE_UNIT_PRICE_CELL]
    if not (isinstance(fpe_price_cell.value, str) and fpe_price_cell.value.startswith("=")):
        fpe_price_cell.value = f"={PRICE_DB_SHEET}!{FPE_MARGIN_CELL}"

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    date_str = order_date.isoformat() if order_date else "unknown-date"
    output_path = output_dir / f"견적서_{_safe_filename_part(model_name)}_{date_str}.xlsx"
    wb.save(output_path)

    return output_path


if __name__ == "__main__":
    order_file = sys.argv[1] if len(sys.argv) > 1 else "inputs/order sheet.xlsx"
    product_type = sys.argv[2] if len(sys.argv) > 2 else input("TYPE을 입력하세요 (예: Interposer): ").strip()
    count_arg = sys.argv[3] if len(sys.argv) > 3 else input("제품 개취수를 입력하세요 (예: 500): ").strip()
    tool_arg = sys.argv[4] if len(sys.argv) > 4 else input("툴비를 포함할까요? (y/n, 기본 y): ").strip()

    try:
        product_count = int(count_arg)
    except ValueError:
        print(f"에러: 제품 개취수는 숫자여야 합니다 - {count_arg!r}")
        sys.exit(1)

    include_tool_charge = tool_arg.strip().lower() not in ("n", "no")

    try:
        saved_path = generate_quote(order_file, product_type, product_count, include_tool_charge)
    except (OrderFileError, QuoteInputError) as e:
        print(f"에러: {e}")
        sys.exit(1)

    print(f"견적서 저장됨: {saved_path}")
    print("엑셀에서 열면 단가 등 계산 항목이 자동으로 채워져 보입니다.")
